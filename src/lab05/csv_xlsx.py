import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    csv_file=Path(csv_path)
    if not csv_file.exists():
        raise FileNotFoundError("Файл не найден")
    if csv_file.suffix != '.csv':
        raise ValueError("Неверный тип файла")
    
    
    #создание excel книги
    wb=Workbook() #создаем новую книгу эксель
    ws=wb.active #получаем активный лист(по умолчанию создаётся один)
    ws.title="Sheet1"#переименовали лист

    with open(csv_path, 'r', encoding='utf-8') as f:
        reader= csv.DictReader(f) #читает сsv как словари (первая строка автоматич - ключи словаря)
        rows = list(reader)
    if len(rows)==0:
        raise ValueError("Файл не содержит данных")
    if not reader.fieldnames: #проверяем наличие заголовков(fildnames содержит названия колонок из первой стр)
        raise ValueError("Файл не содержит заголовка")
    
    ws.append(reader.fieldnames) #записываем заголовки в первою строку таблицы

    r_count=0 #счетчик, чтобы проверить если ли данные, кроме заголовков
    for row in rows:
        r_count+=1

        data_for_ex=[] #то, что буду добавлять в эксель
        for title in reader.fieldnames:
            data_for_ex.append(row[title]) #добавляю значения в правильном порядке(как заголовки)
        ws.append(data_for_ex)
    if r_count == 0:
        raise ValueError("Нет данных")
    

    for col_index in range(1,len(reader.fieldnames)+1): #собираем индексы колонок, начиная с 1
        column_letter=get_column_letter(col_index) #ф-ция, преобразует числовой индекс колонки в букву (как в эксель)
        max_len=0

        #находим макс длину слова в колонке
        for row in ws[column_letter]:
            if row.value is not None: #проверка существования значения в ячейке
                max_len=max(max_len,len(str(row.value)))
        #устанавливаем ширину(не менее 8 символов)
        m_width=max(max_len+2, 8)
        ws.column_dimensions[column_letter].width =m_width #это встроенное свойство объекта Worksheet в библиотеке(предоставляет доступ к управлению колонками листа эксель)


    xlsx_path = Path(xlsx_path)
    wb.save(xlsx_path)

csv_to_xlsx("data/samples/cities.csv","data/out/people.xlsx")